"""Produit le gabarit d'import des invités (EX-ADM-05).

    python exemples/gabarit_invites.py            # écrit à côté de ce fichier
    python exemples/gabarit_invites.py chemin.xlsx

**Pourquoi le script est versionné et pas seulement le classeur.** Un `.xlsx`
est un binaire : personne ne peut lire dans un diff qu'une colonne a changé de
nom. Or `COLONNES` ci-dessous est la définition **normative** de ce que l'import
attend (EX-ADM-05) — renommer « Prénom » ici sans toucher à l'import casserait
l'import en silence. `test_hygiene.py` contrôle cette liste contre la
spécification, ce qu'aucun contrôle ne pourrait faire sur le binaire.

Le classeur produit est versionné à côté, pour n'avoir pas à le régénérer à
chaque fois qu'on veut le remplir.

`openpyxl` n'est PAS dans `requirements.txt` : ce script ne tourne jamais en
production, il se lance à la main quand on refait le gabarit. La dépendance
arrivera avec l'import lui-même.
"""

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

POLICE = "Arial"
ENTETE_FOND = PatternFill("solid", fgColor="2F3E46")
ENTETE_TEXTE = Font(name=POLICE, size=11, bold=True, color="FFFFFF")
EXEMPLE_FOND = PatternFill("solid", fgColor="FFF6D9")
EXEMPLE_TEXTE = Font(name=POLICE, size=10, italic=True, color="8A6D1F")
NORMAL = Font(name=POLICE, size=11)
TITRE = Font(name=POLICE, size=13, bold=True, color="2F3E46")
DISCRET = Font(name=POLICE, size=10, color="6B6259")
FIN = Side(style="thin", color="C9C2B8")
CADRE = Border(left=FIN, right=FIN, top=FIN, bottom=FIN)

COLONNES = [
    ("Identifiant", 16,
     "FACULTATIF. Ne le remplir que pour distinguer deux personnes de même "
     "prénom ET même nom — deux « Marie Meyer », par exemple. C'est la seule "
     "chose qui les sépare : sans lui, l'import les prendrait pour la même "
     "personne et refuserait le fichier (EX-ADM-15).\n\n"
     "Une fois posé, NE PLUS LE CHANGER : c'est lui qui reconnaît la personne "
     "d'un import à l'autre (EX-ADM-13)."),
    ("Table", 14,
     "Le regroupement physique. C'est un simple ATTRIBUT : déplacer quelqu'un "
     "d'une table à l'autre ne crée jamais une seconde personne (EX-ADM-14).\n\n"
     "Peut rester vide."),
    ("Prénom", 18,
     "La casse n'a pas d'importance : « JEAN-PIERRE », « jean-pierre » et "
     "« Jean-Pierre » donnent le même résultat. L'application capitalise une "
     "fois pour toutes, particules et prénoms composés compris (EX-AUTH-21)."),
    ("Nom", 20,
     "Idem : « de rham » deviendra « de Rham », « GAGNEBIN » deviendra "
     "« Gagnebin » (EX-AUTH-21)."),
    ("Genre", 12,
     "H, F, ou VIDE.\n\n"
     "Vide ne veut pas dire « sans genre » : cela veut dire « au choix du "
     "modèle », qui tranchera lui-même (EX-IA-36, EX-IA-37). C'est une case à "
     "remplir quand on la connaît — et on la connaît presque toujours — pour "
     "éviter d'avoir à la demander à quelqu'un à 22 h."),
    ("Responsable", 14,
     "oui / non, ou vide pour non.\n\n"
     "Le responsable de table reçoit un rôle particulier plus tard dans la "
     "soirée. Un seul par table suffit."),
    ("Marié", 12,
     "oui / non, ou vide pour non.\n\n"
     "Réservé aux deux mariés. Ils ont leur propre parcours (EX-AUTH-20)."),
]

EXEMPLES = [
    ["", "1", "jean-pierre", "gagnebin", "H", "oui", ""],
    ["", "1", "marie-josé", "de rham", "F", "", ""],
    ["MEYER-01", "2", "marie", "meyer", "F", "", ""],
    ["MEYER-02", "5", "marie", "meyer", "F", "", ""],
    ["", "", "olivier", "D'ALEMBERT", "", "", ""],
]


def construire(chemin: str) -> None:
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Invités"

    feuille["A1"] = "Liste des invités — à importer dans Le Livre des Convoqués"
    feuille["A1"].font = TITRE
    feuille["A2"] = (
        "Remplacer les cinq lignes d'exemple sur fond crème par la vraie liste. "
        "Ne pas renommer ni déplacer les colonnes. Survoler un titre de colonne "
        "pour lire ce qu'elle attend."
    )
    feuille["A2"].font = DISCRET
    feuille["A3"] = (
        "L'import se fait toujours en simulation d'abord : il montre ce qu'il "
        "ferait avant de le faire, et réimporter le même fichier ne duplique rien."
    )
    feuille["A3"].font = DISCRET

    ligne_entete = 5
    for index, (titre, largeur, aide) in enumerate(COLONNES, start=1):
        cellule = feuille.cell(row=ligne_entete, column=index, value=titre)
        cellule.font = ENTETE_TEXTE
        cellule.fill = ENTETE_FOND
        cellule.border = CADRE
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        commentaire = Comment(aide, "Le Livre des Convoqués")
        commentaire.width, commentaire.height = 330, 190
        cellule.comment = commentaire
        feuille.column_dimensions[get_column_letter(index)].width = largeur
    feuille.row_dimensions[ligne_entete].height = 24

    for décalage, exemple in enumerate(EXEMPLES):
        ligne = ligne_entete + 1 + décalage
        for index, valeur in enumerate(exemple, start=1):
            cellule = feuille.cell(row=ligne, column=index, value=valeur)
            cellule.font = EXEMPLE_TEXTE
            cellule.fill = EXEMPLE_FOND
            cellule.border = CADRE

    premiere_vraie = ligne_entete + 1 + len(EXEMPLES)
    for ligne in range(premiere_vraie, premiere_vraie + 200):
        for index in range(1, len(COLONNES) + 1):
            cellule = feuille.cell(row=ligne, column=index)
            cellule.font = NORMAL
            cellule.border = CADRE

    derniere = premiere_vraie + 199
    genre = DataValidation(type="list", formula1='"H,F"', allow_blank=True,
                           showErrorMessage=False)
    genre.prompt = "H, F, ou laisser vide — vide = au choix du modèle"
    genre.promptTitle = "Genre"
    genre.showInputMessage = True
    feuille.add_data_validation(genre)
    genre.add(f"E{ligne_entete + 1}:E{derniere}")

    for colonne in ("F", "G"):
        oui_non = DataValidation(type="list", formula1='"oui,non"',
                                 allow_blank=True, showErrorMessage=False)
        oui_non.prompt = "oui, non, ou laisser vide (= non)"
        oui_non.promptTitle = "oui / non"
        oui_non.showInputMessage = True
        feuille.add_data_validation(oui_non)
        oui_non.add(f"{colonne}{ligne_entete + 1}:{colonne}{derniere}")

    feuille.freeze_panes = f"A{ligne_entete + 1}"

    aide = classeur.create_sheet("Comment ça marche")
    aide.column_dimensions["A"].width = 26
    aide.column_dimensions["B"].width = 92
    aide["A1"] = "Ce qu'il faut savoir avant d'importer"
    aide["A1"].font = TITRE

    texte = [
        ("Ce que l'import fait",
         "Il crée les personnes absentes, met à jour celles qui ont changé, et "
         "ne touche pas aux autres. Il ne crée AUCUN personnage : les invités "
         "répondent eux-mêmes au questionnaire le soir venu."),
        ("Réimporter est sans risque",
         "Le même fichier importé deux fois ne duplique rien (EX-ADM-06). "
         "Corriger une faute de frappe et réimporter est le geste normal."),
        ("Comment une personne est reconnue",
         "Par la colonne Identifiant si elle est remplie ; sinon par le couple "
         "(Prénom, Nom), accents et casse ignorés (EX-ADM-13)."),
        ("Deux personnes de même nom",
         "Deux « Marie Meyer » sans Identifiant font REFUSER tout le fichier, "
         "avec un rapport nommant les lignes en conflit (EX-ADM-15). Leur "
         "donner à chacune un Identifiant distinct — n'importe lequel, du "
         "moment qu'il ne change plus — les sépare."),
        ("Changer quelqu'un de table",
         "Il suffit de modifier la colonne Table et de réimporter. La table "
         "est un attribut, jamais une clé (EX-ADM-14)."),
        ("Retirer quelqu'un",
         "Une personne retirée du fichier est marquée inactive, jamais "
         "supprimée (EX-ADM-07). Ce qu'elle a écrit n'est jamais perdu."),
        ("La colonne Genre",
         "H, F, ou vide. Vide signifie « au choix du modèle » et non « sans "
         "genre » : un portrait français non genré est illisible (EX-IA-37). "
         "La remplir évite d'avoir à poser la question à quelqu'un à 22 h."),
        ("Ce qui n'est PAS dans ce fichier",
         "Les noms des dix régions ne passent pas par l'import (EX-ADM-23) : "
         "ils se saisissent une fois dans l'administration, et restent "
         "modifiables même en pleine soirée."),
        ("Quelqu'un d'oublié",
         "Ce n'est pas grave. Un invité absent de la liste peut saisir son nom "
         "lui-même, par l'option « Je ne suis pas dans la liste » "
         "(EX-AUTH-19). La liste est un confort, pas une barrière."),
    ]
    for décalage, (titre, corps) in enumerate(texte):
        ligne = 3 + décalage * 2
        aide.cell(row=ligne, column=1, value=titre).font = Font(
            name=POLICE, size=11, bold=True, color="2F3E46")
        cellule = aide.cell(row=ligne, column=2, value=corps)
        cellule.font = NORMAL
        cellule.alignment = Alignment(wrap_text=True, vertical="top")
        aide.row_dimensions[ligne].height = 34

    classeur.save(chemin)


if __name__ == "__main__":
    import pathlib
    import sys

    defaut = pathlib.Path(__file__).parent / "invites-gabarit.xlsx"
    chemin = sys.argv[1] if len(sys.argv) > 1 else str(defaut)
    construire(chemin)
    print(f"gabarit écrit : {chemin}")
