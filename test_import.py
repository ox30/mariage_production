"""Tests de fumée de l'import des invités. Lancer : python test_import.py"""
import base64, os, pathlib, sys
sys.path.insert(0, str(pathlib.Path(__file__).parent))
os.environ["MOT_DE_PASSE_ADMIN"] = "secret"
os.environ.setdefault("WORKER_ACTIF", "0")
os.environ.setdefault("INSTANTANE_ACTIF", "0")
import html as _html
from openpyxl import Workbook
import base_donnees as bd, import_invites, main, modeles, test_outils
from sqlalchemy import select

AUTH = {"Authorization": "Basic " + base64.b64encode(b"a:secret").decode()}
ATELIER = pathlib.Path("essais-import")
ATELIER.mkdir(exist_ok=True)
c = test_outils.client(main.app)


def classeur(lignes, nom="essai.xlsx", entete=import_invites.COLONNES):
    """Écrit un classeur minimal. L'en-tête est un paramètre : un fichier au
    format cassé doit être éprouvé, pas seulement le format juste."""
    livre = Workbook()
    feuille = livre.active
    feuille.append(entete)
    for ligne in lignes:
        feuille.append(list(ligne))
    chemin = ATELIER / nom
    livre.save(chemin)
    return chemin


def actives():
    with bd.Seance() as seance:
        return list(seance.scalars(
            select(modeles.Personne).where(modeles.Personne.active.is_(True))))


def par_nom(prenom, nom):
    return bd.resoudre(prenom, nom).candidates


# --- Le gabarit livré se lit, et se capitalise ------------------------------
plan = import_invites.preparer(
    pathlib.Path("exemples") / "invites-gabarit.xlsx")
assert plan.recevable, plan.conflits + plan.erreurs
assert plan.lignes_lues == 5, plan.lignes_lues
# Les deux Marie Meyer du gabarit ne sont PAS en conflit : elles portent des
# identifiants distincts, ce qui est la seule raison d'être de la colonne.
assert not plan.conflits
# Et le gabarit signale ses propres lignes d'exemple : oubliées dans un fichier
# rempli, elles fabriqueraient des invités fictifs au milieu des vrais.
assert len(plan.avertissements) >= 4, plan.avertissements
assert all("exemple" in a for a in plan.avertissements)
print("TOUT PASSE — le gabarit livré se lit et signale ses lignes d'exemple")

# --- L'en-tête est cherché, pas supposé en ligne 1 -------------------------
# Le gabarit porte un titre et deux lignes d'explication au-dessus ; un fichier
# repassé par un tableur peut en gagner d'autres.
livre = Workbook()
f = livre.active
f.append(["Liste des invités — mariage"])
f.append([])
f.append(import_invites.COLONNES)
f.append(["", "3", "aria", "sonval", "F", "non", "non"])
decale = ATELIER / "decale.xlsx"
livre.save(decale)
plan = import_invites.preparer(decale)
assert plan.recevable and plan.lignes_lues == 1, (plan.erreurs, plan.lignes_lues)

# Un fichier sans en-tête reconnaissable le dit, et nomme les colonnes.
casse = classeur([["x"]], "casse.xlsx", entete=["Machin", "Truc"])
plan = import_invites.preparer(casse)
assert not plan.recevable
assert "Prénom" in " ".join(plan.erreurs), plan.erreurs
assert "gabarit" in " ".join(plan.erreurs), "le message doit dire où trouver le modèle"
print("TOUT PASSE — l'en-tête se cherche, et son absence se dit")

# --- La simulation n'écrit rien (EX-ADM-16) -------------------------------
avant = len(actives())
fichier = classeur([
    ["", "1", "jean-pierre", "GAGNEBIN", "H", "oui", "non"],
    ["", "1", "marie-josé", "de rham", "F", "non", "non"],
    ["", "2", "aria", "sonval", "", "non", "non"],
], "trois.xlsx")
plan = import_invites.preparer(fichier)
assert len(plan.creations) == 3, plan.creations
assert len(actives()) == avant, "la simulation a écrit en base"
print("TOUT PASSE — la simulation ne touche pas à la base")

# --- L'application crée, et capitalise (EX-AUTH-21) -----------------------
plan = import_invites.appliquer(fichier)
assert plan.recevable
assert len(actives()) == avant + 3, len(actives())
jp = par_nom("jean-pierre", "gagnebin")[0]
assert (jp.prenom, jp.nom) == ("Jean-Pierre", "Gagnebin"), (jp.prenom, jp.nom)
assert par_nom("marie-josé", "de rham")[0].nom == "de Rham", "particule conservée"
assert jp.genre == "masculin" and jp.est_responsable is True
assert jp.source == "import"
# Vide ne veut pas dire « sans genre » mais « au choix du modèle » (EX-IA-37).
assert par_nom("aria", "sonval")[0].genre is None
print("TOUT PASSE — l'import crée les personnes et capitalise leur nom")

# --- Réimporter le même fichier ne duplique rien (EX-ADM-06) --------------
reference = len(actives())
plan = import_invites.appliquer(fichier)
assert len(actives()) == reference, "l'import a dupliqué"
assert not plan.creations and not plan.modifications, (plan.creations, plan.modifications)
assert len(plan.inchangees) == 3
assert plan.sans_effet, "un réimport identique doit être annoncé sans effet"
print("TOUT PASSE — réimporter le même fichier ne duplique rien")

# --- La table est un attribut, pas une clé (EX-ADM-14) -------------------
uuid_jp = par_nom("jean-pierre", "gagnebin")[0].uuid
deplace = classeur([
    ["", "7", "jean-pierre", "GAGNEBIN", "H", "oui", "non"],
    ["", "1", "marie-josé", "de rham", "F", "non", "non"],
    ["", "2", "aria", "sonval", "", "non", "non"],
], "deplace.xlsx")
plan = import_invites.appliquer(deplace)
assert len(actives()) == reference, "déplacer quelqu'un a créé une seconde personne"
assert par_nom("jean-pierre", "gagnebin")[0].uuid == uuid_jp, "l'uuid a changé"
with bd.Seance() as seance:
    personne = seance.get(modeles.Personne, uuid_jp)
    assert seance.get(modeles.TableGroupe, personne.table_uuid).nom == "7"
print("TOUT PASSE — changer de table ne crée pas une seconde personne")

# --- Deux lignes de même clé rejettent le fichier ENTIER (EX-ADM-15) -----
avant_conflit = len(actives())
double = classeur([
    ["", "1", "Marie", "Meyer", "F", "non", "non"],
    ["", "5", "marie", "MEYER", "F", "non", "non"],
    ["", "3", "Nouvelle", "Personne", "F", "non", "non"],
], "double.xlsx")
plan = import_invites.preparer(double)
assert not plan.recevable, "deux homonymes sans identifiant doivent rejeter"
assert len(plan.conflits) == 1, plan.conflits
# Numéros DÉRIVÉS du classeur, pas écrits en dur : mon premier jet attendait
# 1 et 3 pour un fichier dont l'en-tête occupe la ligne 1, et l'échec ne disait
# rien du défaut réel. Même faute que sur les doublons de config.yaml.
from openpyxl import load_workbook as _lire
_f = _lire(double)[_lire(double).sheetnames[0]]
_meyer = [n for n, r in enumerate(_f.iter_rows(values_only=True), start=1)
          if r and import_invites.normaliser(str(r[3] or "")) == "meyer"]
assert len(_meyer) == 2, _meyer
for _n in _meyer:
    assert str(_n) in plan.conflits[0], \
        f"la ligne {_n} n'est pas nommée : {plan.conflits[0]}"
plan = import_invites.appliquer(double)
assert len(actives()) == avant_conflit, \
    "une ligne saine a été importée malgré le rejet — l'état devient indéchiffrable"
assert not par_nom("Nouvelle", "Personne"), "rejet partiel"
print("TOUT PASSE — deux lignes de même clé rejettent le fichier entier")

# --- L'identifiant distingue deux homonymes (EX-ADM-13) -----------------
meyer = classeur([
    ["MEYER-01", "1", "Marie", "Meyer", "F", "non", "non"],
    ["MEYER-02", "5", "marie", "MEYER", "F", "non", "non"],
], "meyer.xlsx")
plan = import_invites.appliquer(meyer)
assert plan.recevable, plan.conflits
deux = par_nom("Marie", "Meyer")
assert len(deux) == 2, f"{len(deux)} Marie Meyer au lieu de deux"
assert {p.identifiant_import for p in deux} == {"MEYER-01", "MEYER-02"}

# Et réimporter ne les fusionne pas : c'est l'identifiant qui les rattrape,
# pas leur nom. Un rapprochement par nom en désignerait une au hasard.
import_invites.appliquer(meyer)
assert len(par_nom("Marie", "Meyer")) == 2, "les homonymes ont fusionné au réimport"

# Déplacer la seconde ne touche pas la première.
uuids_avant = {p.identifiant_import: p.uuid for p in par_nom("Marie", "Meyer")}
import_invites.appliquer(classeur([
    ["MEYER-01", "1", "Marie", "Meyer", "F", "non", "non"],
    ["MEYER-02", "9", "marie", "MEYER", "F", "oui", "non"],
], "meyer2.xlsx"))
apres = {p.identifiant_import: p for p in par_nom("Marie", "Meyer")}
assert {k: v.uuid for k, v in apres.items()} == uuids_avant, "un uuid a bougé"
assert apres["MEYER-02"].est_responsable is True
assert apres["MEYER-01"].est_responsable is False, "la mauvaise des deux a été modifiée"
print("TOUT PASSE — l'Identifiant distingue deux homonymes, d'un import à l'autre")

# --- Le genre s'enrichit, ne s'efface jamais (EX-IA-37) -----------------
import_invites.appliquer(classeur(
    [["", "2", "aria", "sonval", "F", "non", "non"]], "genre1.xlsx"))
assert par_nom("aria", "sonval")[0].genre == "feminin"
import_invites.appliquer(classeur(
    [["", "2", "aria", "sonval", "", "non", "non"]], "genre2.xlsx"))
assert par_nom("aria", "sonval")[0].genre == "feminin", \
    "une colonne vide veut dire « au choix du modèle », pas « oublie ce que tu sais »"

# Une valeur incomprise est refusée, elle ne devient pas « vide » en silence.
plan = import_invites.preparer(classeur(
    [["", "2", "x", "y", "Madame", "non", "non"]], "genre3.xlsx"))
assert not plan.recevable and "Madame" in " ".join(plan.erreurs), plan.erreurs
print("TOUT PASSE — le genre s'enrichit sans jamais s'effacer")

# --- Sans la case « liste complète », personne n'est désactivé ----------
partiel = classeur([["", "1", "jean-pierre", "GAGNEBIN", "H", "oui", "non"]],
                   "partiel.xlsx")
combien = len(actives())
plan = import_invites.preparer(partiel)
assert not plan.inactivations, \
    "un import partiel désactiverait tout le monde en silence"
import_invites.appliquer(partiel)
assert len(actives()) == combien, "des personnes ont été désactivées sans la case"
print("TOUT PASSE — sans « liste complète », l'import n'est jamais soustractif")

# --- Avec la case, l'absent est désactivé — sauf s'il a joué (EX-ADM-07) --
uuid_aria = par_nom("aria", "sonval")[0].uuid
chronique_aria = bd.creer(uuid_aria, {"metier": "luthière"}, main.CODES_LIEUX)
assert chronique_aria

plan = import_invites.preparer(partiel, liste_complete=True)
noms_desactives = {f"{p.prenom} {p.nom}" for p in plan.inactivations}
noms_proteges = {f"{p.prenom} {p.nom}" for p in plan.protegees}
assert "Aria Sonval" in noms_proteges, \
    "une personne ayant écrit son personnage ne doit jamais être désactivée"
assert "Aria Sonval" not in noms_desactives
assert "Marie Meyer" in noms_desactives, noms_desactives

import_invites.appliquer(partiel, liste_complete=True)
# EX-ADM-07 — inactive, jamais supprimée.
with bd.Seance() as seance:
    total = seance.scalar(select(modeles.Personne).where(
        modeles.Personne.prenom == "Marie"))
    assert total is not None, "une personne a été SUPPRIMÉE au lieu d'être inactivée"

# Et celle qui a joué reste trouvable : la désactiver l'aurait retirée de la
# recherche par nom, et elle n'aurait plus pu revoir son propre personnage.
assert par_nom("aria", "sonval"), "la personne protégée est devenue introuvable"
assert bd.chronique_de_personne(uuid_aria) == chronique_aria
print("TOUT PASSE — l'absent est désactivé, jamais celui qui a déjà joué")

# --- Une désactivée peut revenir ----------------------------------------
import_invites.appliquer(classeur([
    ["", "1", "jean-pierre", "GAGNEBIN", "H", "oui", "non"],
    ["MEYER-01", "1", "Marie", "Meyer", "F", "non", "non"],
], "retour.xlsx"))
assert par_nom("Marie", "Meyer"), "une personne réintégrée reste invisible"
print("TOUT PASSE — une personne retirée puis réintégrée redevient active")

# --- L'écran d'administration a sa propre porte -------------------------
assert c.get("/admin/invites").status_code == 401, "l'import est ouvert à tous"
assert c.post("/admin/invites/simuler").status_code == 401
assert c.post("/admin/invites/appliquer").status_code == 401
assert c.get("/admin/invites", headers=AUTH).status_code == 200

# Et le nom de fichier de la confirmation ne sort pas du dossier d'imports.
# Le piège est un classeur RÉELLEMENT atteignable par traversée : mon premier
# jet visait « ../../config.yaml », qui n'existe pas à cet endroit — le test
# passait donc grâce à l'absence de la cible, pas grâce au filtre.
import config as _config

# Le dossier d'imports n'existe qu'après le premier envoi réussi ; le créer
# ici, sinon le noyau ne résout pas le « .. » et le contrôle du piège échoue
# pour une raison qui n'a rien à voir avec ce qu'il éprouve.
(_config.projet().dossier / "imports").mkdir(parents=True, exist_ok=True)
_piege = _config.projet().dossier / "piege.xlsx"
_source = classeur([["", "1", "Cheval", "DeTroie", "H", "non", "non"]], "piege.xlsx")
_piege.write_bytes(_source.read_bytes())
assert (_config.projet().dossier / "imports" / ".." / "piege.xlsx").is_file(), \
    "le piège doit être atteignable par traversée, sinon le test ne prouve rien"

r = c.post("/admin/invites/appliquer", headers=AUTH,
           data={"fichier": "../piege.xlsx"}, follow_redirects=False)
assert r.status_code == 303 and r.headers["location"] == "/admin/invites", r.headers
assert not par_nom("Cheval", "DeTroie"), \
    "un fichier hors du dossier d'imports a été lu et appliqué"
_piege.unlink()
print("TOUT PASSE — l'écran d'import est fermé, et son nom de fichier borné")

# --- Le parcours complet par l'écran ------------------------------------
with open(classeur([["", "4", "bilbon", "sacquet", "H", "non", "non"]],
                   "ecran.xlsx"), "rb") as fichier_ouvert:
    r = c.post("/admin/invites/simuler", headers=AUTH,
               files={"classeur": ("liste.xlsx", fichier_ouvert.read())})
texte = _html.unescape(r.text)
assert "Simulation" in texte and "rien n'est encore écrit" in texte, texte[:400]
assert not par_nom("bilbon", "sacquet"), "la simulation a écrit"

jeton = texte.split('name="fichier" value="')[1].split('"')[0]
r = c.post("/admin/invites/appliquer", headers=AUTH, data={"fichier": jeton})
assert "Import effectué" in _html.unescape(r.text), r.text[:400]
assert par_nom("bilbon", "sacquet")[0].prenom == "Bilbon"

# Le classeur est conservé sur le volume : la confirmation le relit, il part
# dans les sauvegardes, et l'on sait ce qui a été importé.
import config
conserves = list((config.projet().dossier / "imports").glob("*.xlsx"))
assert conserves, "aucun classeur conservé"
print("TOUT PASSE — l'écran enchaîne simulation puis application")

import shutil
shutil.rmtree(ATELIER, ignore_errors=True)
