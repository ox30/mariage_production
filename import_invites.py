"""L'import de la liste des invités (EX-ADM-05, EX-ADM-06, EX-ADM-13 à 16).

Le module ne décide rien tout seul : il lit le classeur, en tire un **plan**, et
n'écrit que si on le lui demande explicitement. C'est `EX-ADM-16` — la
simulation d'abord — et c'est aussi le seul garde-fou possible contre un fichier
préparé la veille à minuit et relu par personne.

Quatre propriétés, chacune apprise d'un défaut :

- **Idempotent** (`EX-ADM-06`). Réimporter le même fichier ne duplique rien.
  Corriger une faute de frappe et réimporter est le geste normal, pas une
  manœuvre risquée.
- **Clé de rapprochement** (`EX-ADM-13`). `Identifiant` s'il est rempli, sinon
  le couple (Prénom, Nom) normalisé.
- **La table est un attribut** (`EX-ADM-14`), jamais une clé : déplacer
  quelqu'un ne crée pas une seconde personne.
- **Deux lignes de même clé font rejeter le fichier entier** (`EX-ADM-15`),
  avec un rapport nommant les lignes. Importer à moitié serait pire : on ne
  saurait plus ce qui est passé.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field

from sqlalchemy import select

from base_donnees import Seance, journaliser
from modeles import Chronique, Personne, TableGroupe

# EX-ADM-05 — l'ordre fait foi. `exemples/gabarit_invites.py` produit ces
# colonnes et `test_hygiene.py` vérifie qu'il ne s'en écarte pas.
COLONNES = ["Identifiant", "Table", "Prénom", "Nom", "Genre",
            "Responsable", "Marié"]

# Une ligne d'exemple du gabarit oubliée dans le fichier rempli produirait un
# invité fictif au milieu des vrais.
#
# La détection porte sur la ligne ENTIÈRE, et non sur le seul couple
# (prénom, nom). Premier jet le 25 août : la vraie liste contenait un
# « Jean-Pierre Gagnebin » — mes noms d'exemple venaient des tests du projet,
# donc de vrais invités — et le rapport accusait quelqu'un d'être fictif. Un
# garde-fou qui crie au loup sur un vrai invité use la confiance qu'on lui
# porte, et c'est le jour où il aura raison qu'on ne le lira plus.
#
# Comparer table, genre et rôle en plus du nom rend la coïncidence
# invraisemblable : l'exemple est à la table 1 et responsable ; le vrai
# Jean-Pierre est à la table 2 et ne l'est pas.
EXEMPLES_DU_GABARIT = {
    ("1", "jean-pierre", "gagnebin", "h", "oui", ""),
    ("1", "marie-jose", "de rham", "f", "", ""),
    ("2", "marie", "meyer", "f", "", ""),
    ("5", "marie", "meyer", "f", "", ""),
    ("", "olivier", "d'alembert", "", "", ""),
}

VRAI = {"oui", "o", "yes", "y", "1", "true", "vrai", "x"}
FAUX = {"", "non", "n", "no", "0", "false", "faux"}
GENRES = {"h": "masculin", "m": "masculin", "homme": "masculin",
          "f": "feminin", "femme": "feminin"}


def normaliser(valeur: str) -> str:
    """Casse, accents et espaces retirés — la forme qui sert de clé.

    Les accents sont dépouillés parce qu'une liste tapée sur trois claviers
    différents contient « Gaëlle » et « Gaelle », et que ce sont la même
    personne. La forme normalisée ne sert qu'à comparer : ce qui s'affiche
    reste ce qui a été écrit.
    """
    depouille = unicodedata.normalize("NFKD", str(valeur or "").strip().lower())
    depouille = "".join(c for c in depouille if not unicodedata.combining(c))
    return " ".join(depouille.split())


def _booleen(valeur) -> bool:
    brut = normaliser(valeur)
    if brut in VRAI:
        return True
    if brut in FAUX:
        return False
    raise ValeurRefusee(f"« {valeur} » n'est ni oui ni non")


def _genre(valeur) -> str | None:
    """H, F, ou vide. Vide veut dire « au choix du modèle » (EX-IA-37).

    Et non « sans genre » : un portrait français non genré est illisible. La
    distinction compte, parce qu'une colonne laissée vide par négligence et une
    colonne laissée vide à dessein produisent le même résultat — c'est pourquoi
    le rapport compte les vides séparément.
    """
    brut = normaliser(valeur)
    if not brut:
        return None
    if brut in GENRES:
        return GENRES[brut]
    raise ValeurRefusee(f"« {valeur} » n'est ni H ni F")


class ValeurRefusee(ValueError):
    pass


def _lister(valeurs, maximum: int = 12) -> str:
    """Une énumération lisible, tronquée plutôt qu'interminable."""
    valeurs = list(valeurs)
    debut = ", ".join(str(v) for v in valeurs[:maximum])
    reste = len(valeurs) - maximum
    return f"{debut} et {reste} autre(s)" if reste > 0 else debut


def _grouper(motif: str, numeros: list[int]) -> str:
    if len(numeros) == 1:
        return f"ligne {numeros[0]} : {motif}"
    return f"{len(numeros)} lignes — {motif} : {_lister(numeros)}"


@dataclass
class Ligne:
    numero: int
    identifiant: str
    table: str
    prenom: str
    nom: str
    genre: str | None
    est_responsable: bool
    est_marie: bool

    @property
    def empreinte_exemple(self) -> tuple[str, ...]:
        """La ligne entière, pour reconnaître un exemple NON MODIFIÉ."""
        return (normaliser(self.table), normaliser(self.prenom),
                normaliser(self.nom),
                {"masculin": "h", "feminin": "f"}.get(self.genre, ""),
                "oui" if self.est_responsable else "",
                "oui" if self.est_marie else "")

    @property
    def cle(self) -> tuple[str, ...]:
        """EX-ADM-13 — l'identifiant prime, le couple normalisé sinon."""
        if self.identifiant:
            return ("id", normaliser(self.identifiant))
        return ("nom", normaliser(self.prenom), normaliser(self.nom))


@dataclass
class Plan:
    """Ce que l'import ferait. Rien n'est écrit tant qu'on ne l'applique pas."""

    creations: list[Ligne] = field(default_factory=list)
    modifications: list[tuple[Ligne, str, list[str]]] = field(default_factory=list)
    inchangees: list[Ligne] = field(default_factory=list)
    inactivations: list[Personne] = field(default_factory=list)
    protegees: list[Personne] = field(default_factory=list)
    conflits: list[str] = field(default_factory=list)
    erreurs: list[str] = field(default_factory=list)
    avertissements: list[str] = field(default_factory=list)
    lignes_lues: int = 0

    @property
    def recevable(self) -> bool:
        return not self.conflits and not self.erreurs

    @property
    def sans_effet(self) -> bool:
        return not (self.creations or self.modifications or self.inactivations)


def lire_classeur(chemin) -> tuple[list[Ligne], list[str]]:
    """Lit le fichier et rend les lignes, plus les erreurs de forme.

    Une erreur de forme n'interrompt pas la lecture : rendre les vingt erreurs
    d'un coup vaut mieux que d'en donner une, se faire corriger, et en donner
    une autre. Le fichier est de toute façon rejeté en bloc.
    """
    from openpyxl import load_workbook

    classeur = load_workbook(chemin, data_only=True, read_only=True)
    feuille = classeur[classeur.sheetnames[0]]

    lignes, erreurs = [], []
    groupes: dict[str, list[int]] = {}
    entete_vue, index = None, {}
    for numero, brut in enumerate(feuille.iter_rows(values_only=True), start=1):
        valeurs = ["" if v is None else str(v).strip() for v in brut]
        if not any(valeurs):
            continue

        if entete_vue is None:
            # L'en-tête est cherché plutôt que supposé en ligne 1 : le gabarit
            # porte un titre et deux lignes d'explication au-dessus, et un
            # fichier repassé par un tableur peut en gagner d'autres.
            normalisees = [normaliser(v) for v in valeurs]
            if all(normaliser(c) in normalisees for c in COLONNES):
                entete_vue = numero
                index = {c: normalisees.index(normaliser(c)) for c in COLONNES}
            continue

        def champ(nom_colonne: str) -> str:
            position = index[nom_colonne]
            return valeurs[position] if position < len(valeurs) else ""

        if not champ("Prénom").strip() and not champ("Nom").strip():
            continue
        try:
            # Le NOM DE FAMILLE est facultatif, le prénom ne l'est pas.
            #
            # Sur la vraie liste du 25 août, 48 invités sur 93 n'en avaient pas
            # : le conjoint d'un cousin, l'ami d'enfance dont on n'a jamais su
            # le nom. Les exiger tous rejetait la moitié du fichier et forçait
            # à inventer. Le rapprochement reste sûr — la clé devient
            # (prénom, ""), et deux « Sophie » sans nom déclenchent le conflit
            # d'EX-ADM-15 comme deux homonymes ordinaires.
            if not champ("Prénom").strip():
                raise ValeurRefusee("aucun prénom")
            lignes.append(Ligne(
                numero=numero,
                identifiant=champ("Identifiant").strip(),
                table=champ("Table").strip(),
                prenom=champ("Prénom").strip(),
                nom=champ("Nom").strip(),
                genre=_genre(champ("Genre")),
                est_responsable=_booleen(champ("Responsable")),
                est_marie=_booleen(champ("Marié")),
            ))
        except ValeurRefusee as exc:
            # Groupées par NATURE et non ligne à ligne : le rapport du 25 août
            # alignait quarante-huit fois le même message, ce qui ne se lit
            # pas et cache les autres erreurs sous la répétition.
            groupes.setdefault(str(exc), []).append(numero)

    classeur.close()
    for motif, numeros in groupes.items():
        erreurs.append(_grouper(motif, numeros))
    if entete_vue is None:
        erreurs.insert(0, (
            "aucune ligne d'en-tête trouvée. Le fichier doit porter les sept "
            f"colonnes d'EX-ADM-05 sur une même ligne : {', '.join(COLONNES)}. "
            "Le gabarit est dans exemples/invites-gabarit.xlsx."))
    return lignes, erreurs


def preparer(chemin, liste_complete: bool = False) -> Plan:
    """Calcule le plan sans rien écrire (EX-ADM-16)."""
    lignes, erreurs = lire_classeur(chemin)
    plan = Plan(erreurs=erreurs, lignes_lues=len(lignes))

    # EX-ADM-15 — deux lignes de même clé rejettent le fichier ENTIER. Importer
    # les autres laisserait un état dont personne ne saurait dire ce qu'il
    # contient.
    vues: dict[tuple, int] = {}
    for ligne in lignes:
        if ligne.cle in vues:
            quoi = ("le même Identifiant" if ligne.cle[0] == "id"
                    else "le même prénom et le même nom, sans Identifiant "
                         "pour les distinguer")
            plan.conflits.append(
                f"lignes {vues[ligne.cle]} et {ligne.numero} : "
                f"{ligne.prenom} {ligne.nom} — {quoi}")
        else:
            vues[ligne.cle] = ligne.numero

    sans_nom = [l for l in lignes if not l.nom.strip()]
    if sans_nom:
        plan.avertissements.append(
            f"{len(sans_nom)} personne(s) sans nom de famille : "
            f"{_lister(l.prenom for l in sans_nom)}. Elles seront importées, "
            "mais leur palier d'indice ne montrera qu'une initiale, et deux "
            "personnes du même prénom devront être distinguées.")

    for ligne in lignes:
        if ligne.empreinte_exemple in EXEMPLES_DU_GABARIT:
            plan.avertissements.append(
                f"ligne {ligne.numero} : « {ligne.prenom} {ligne.nom} » est une "
                "ligne d'exemple du gabarit. À supprimer si ce n'est pas un "
                "vrai invité.")

    if not plan.recevable:
        return plan

    with Seance() as seance:
        vivantes = list(seance.scalars(select(Personne)))
        par_identifiant = {normaliser(p.identifiant_import): p for p in vivantes
                           if p.identifiant_import}
        par_nom: dict[tuple, list[Personne]] = {}
        for p in vivantes:
            par_nom.setdefault((normaliser(p.prenom), normaliser(p.nom)),
                               []).append(p)

        touchees: set[str] = set()
        for ligne in lignes:
            existante = _retrouver(ligne, par_identifiant, par_nom)
            if existante is None:
                plan.creations.append(ligne)
                continue
            touchees.add(existante.uuid)
            champs = _differences(seance, ligne, existante)
            if champs:
                plan.modifications.append(
                    (ligne, f"{existante.prenom} {existante.nom}", champs))
            else:
                plan.inchangees.append(ligne)

        if liste_complete:
            for personne in vivantes:
                if personne.uuid in touchees or not personne.active:
                    continue
                if personne.source != "import":
                    # Une saisie libre n'a jamais été dans le fichier : le
                    # fichier ne peut donc pas l'en avoir retirée.
                    continue
                a_une_chronique = seance.scalar(
                    select(Chronique.uuid).where(
                        Chronique.personne_uuid == personne.uuid,
                        Chronique.supprimee.is_(False)))
                if a_une_chronique:
                    # Le fichier ne fait pas autorité contre un fait accompli.
                    # L'inactiver la retirerait de `personnes_par_nom`, et elle
                    # ne pourrait plus revoir son propre personnage.
                    plan.protegees.append(personne)
                else:
                    plan.inactivations.append(personne)

    return plan


def _retrouver(ligne, par_identifiant, par_nom) -> Personne | None:
    """EX-ADM-13. L'identifiant prime, y compris sur un homonyme."""
    if ligne.identifiant:
        return par_identifiant.get(normaliser(ligne.identifiant))
    candidates = par_nom.get((normaliser(ligne.prenom), normaliser(ligne.nom)), [])
    # Une personne déjà distinguée par un identifiant ne se rattrape pas par son
    # nom : sinon une ligne sans identifiant écraserait l'une des deux Marie
    # Meyer, au hasard de l'ordre de lecture.
    libres = [p for p in candidates if not p.identifiant_import]
    return libres[0] if len(libres) == 1 else None


def _differences(seance, ligne: Ligne, personne: Personne) -> list[str]:
    champs = []
    if normaliser(personne.prenom) != normaliser(ligne.prenom) or \
            normaliser(personne.nom) != normaliser(ligne.nom):
        champs.append(f"nom : {personne.prenom} {personne.nom} → "
                      f"{ligne.prenom} {ligne.nom}")
    if ligne.genre != personne.genre:
        # Le fichier ne peut qu'ENRICHIR le genre, jamais l'effacer : une
        # colonne laissée vide veut dire « au choix du modèle », pas « oublie
        # ce que tu sais » (EX-IA-37).
        if ligne.genre is not None:
            champs.append(f"genre : {personne.genre or '—'} → {ligne.genre}")
    if ligne.est_responsable != personne.est_responsable:
        champs.append("responsable de table" if ligne.est_responsable
                      else "plus responsable de table")
    if ligne.est_marie != personne.est_marie:
        champs.append("marié·e" if ligne.est_marie else "plus marié·e")
    code_table = _code_de_table(seance, personne.table_uuid)
    if normaliser(code_table) != normaliser(ligne.table):
        # EX-ADM-14 — la table change sans jamais créer une seconde personne.
        champs.append(f"table : {code_table or '—'} → {ligne.table or '—'}")
    if not personne.active:
        champs.append("réactivée")
    return champs


def _code_de_table(seance, table_uuid: str | None) -> str:
    """Le CODE, parce que c'est lui que le fichier porte.

    Comparer le nom affiché ferait apparaître un changement de table à chaque
    simulation dès qu'une table est renommée : « table 3 → 3 », alors que rien
    n'a bougé.
    """
    if not table_uuid:
        return ""
    table = seance.get(TableGroupe, table_uuid)
    return table.code if table else ""


def appliquer(chemin, liste_complete: bool = False) -> Plan:
    """Écrit le plan. Le classeur est **relu** plutôt que le plan mémorisé.

    Un plan calculé il y a dix minutes décrit une base qui a pu changer entre
    temps — quelqu'un a pu créer son personnage pendant la lecture du rapport.
    Relire coûte une seconde et supprime la classe entière des états périmés.
    """
    plan = preparer(chemin, liste_complete=liste_complete)
    if not plan.recevable:
        return plan

    with Seance() as seance:
        # Rapprochées par leur CODE et non par leur nom. La colonne « Table »
        # du fichier porte « 3 » ; le nom affiché peut être devenu
        # « Fondcombe ». Chercher par nom aurait créé une seconde table « 3 »
        # au premier réimport, et y aurait déplacé ses dix invités en silence.
        tables = {normaliser(t.code): t
                  for t in seance.scalars(select(TableGroupe))}

        def table_pour(code: str) -> str | None:
            if not code.strip():
                return None
            existante = tables.get(normaliser(code))
            if existante is None:
                # Une table neuve prend son code comme nom : elle est
                # renommable ensuite dans /admin/tables.
                existante = TableGroupe(code=code.strip(), nom=code.strip(),
                                        ordre=len(tables))
                seance.add(existante)
                seance.flush()
                tables[normaliser(code)] = existante
            return existante.uuid

        vivantes = list(seance.scalars(select(Personne)))
        par_identifiant = {normaliser(p.identifiant_import): p for p in vivantes
                           if p.identifiant_import}
        par_nom: dict[tuple, list[Personne]] = {}
        for p in vivantes:
            par_nom.setdefault((normaliser(p.prenom), normaliser(p.nom)),
                               []).append(p)

        lignes, _ = lire_classeur(chemin)
        for ligne in lignes:
            personne = _retrouver(ligne, par_identifiant, par_nom)
            if personne is None:
                personne = Personne(prenom=_capitaliser(ligne.prenom),
                                    nom=_capitaliser(ligne.nom),
                                    source="import")
                seance.add(personne)
                seance.flush()
            personne.identifiant_import = ligne.identifiant or None
            personne.table_uuid = table_pour(ligne.table)
            personne.est_responsable = ligne.est_responsable
            personne.est_marie = ligne.est_marie
            personne.active = True
            if ligne.genre is not None:
                personne.genre = ligne.genre
            if normaliser(personne.prenom) != normaliser(ligne.prenom):
                personne.prenom = _capitaliser(ligne.prenom)
            if normaliser(personne.nom) != normaliser(ligne.nom):
                personne.nom = _capitaliser(ligne.nom)

        for personne in plan.inactivations:
            attachee = seance.get(Personne, personne.uuid)
            if attachee is not None:
                # EX-ADM-07 — inactive, jamais supprimée. Ce qu'elle a écrit
                # n'est jamais perdu.
                attachee.active = False

        journaliser(seance, "import_invites", objet_type="personne", details={
            "creations": len(plan.creations),
            "modifications": len(plan.modifications),
            "inchangees": len(plan.inchangees),
            "inactivations": len(plan.inactivations),
            "protegees": len(plan.protegees),
            "liste_complete": liste_complete,
        })
        seance.commit()
    return plan


def _capitaliser(valeur: str) -> str:
    """EX-AUTH-21 — particules et prénoms composés, déjà éprouvé par noms.py.

    « de rham » → « de Rham », « JEAN-PIERRE » → « Jean-Pierre ». La liste
    tapée sur trois claviers par quatre personnes n'a aucune chance d'être
    homogène, et c'est ici qu'on la rend présentable une fois pour toutes.
    """
    import noms

    return noms.capitaliser(valeur)
